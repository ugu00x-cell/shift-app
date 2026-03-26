"""
export.py — Excel / CSV エクスポートモジュール
介護シフト自動作成アプリ

生成されたシフトデータを、整形済みの Excel ファイル (.xlsx) または
CSV ファイルとして出力する。
"""

import calendar
import csv
import io
from datetime import date
from io import BytesIO

import jpholiday
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 定数: アサインメント → 日本語表示ラベル
# ---------------------------------------------------------------------------
ASSIGNMENT_LABELS = {
    "day_pattern1":    "デイ8:30-17:30",
    "day_pattern2":    "デイ9:00-16:00",
    "day_pattern3":    "デイ午前のみ",
    "day_pattern4":    "デイ午後のみ",
    "visit_am":        "訪問午前のみ",
    "visit_pm":        "訪問午後のみ",
    "day_p3_visit_pm": "兼務(③→訪問)",
    "visit_am_day_p4": "兼務(訪問→④)",
    "cook_early":      "調理①",
    "cook_morning":    "調理②",
    "cook_late":       "調理③",
    "cook_long":       "調理④",
    # 旧名の後方互換
    "day_am":          "デイ午前のみ",
    "day_pm":          "デイ午後のみ",
    "day_am_visit_pm": "兼務(③→訪問)",
    "visit_am_day_pm": "兼務(訪問→④)",
}

# カテゴリごとの背景色 (アサインメントセル)
ASSIGNMENT_FILL = {
    "day_pattern1":    PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "day_pattern2":    PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid"),
    "day_pattern3":    PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
    "day_pattern4":    PatternFill(start_color="BAE6FD", end_color="BAE6FD", fill_type="solid"),
    "visit_am":        PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "visit_pm":        PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "day_p3_visit_pm": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    "visit_am_day_p4": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    "cook_early":      PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "cook_morning":    PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
    "cook_late":       PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid"),
    "cook_long":       PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid"),
    # 旧名の後方互換
    "day_am":          PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
    "day_pm":          PatternFill(start_color="BAE6FD", end_color="BAE6FD", fill_type="solid"),
    "day_am_visit_pm": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    "visit_am_day_pm": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
}

# 曜日名
WEEKDAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]

# サマリー列ヘッダー (ケア)
SUMMARY_HEADERS = ["デイ午前", "デイ午後", "訪問午前", "訪問午後", "兼務者数", "電話当番"]

# サマリー列ヘッダー (調理)
COOK_SUMMARY_HEADERS = ["調理配置数"]

# ---------------------------------------------------------------------------
# スタイル定義
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="メイリオ", bold=True, size=16)
NORMAL_FONT = Font(name="メイリオ", size=10)
SATURDAY_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
SUNDAY_FILL = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ALERT_FONT = Font(name="メイリオ", size=10, color="CC0000")
WARNING_HEADER_FILL = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
WARNING_HEADER_FONT = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ⑧ 祝日行の背景色
HOLIDAY_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

def _format_break_comment(break_start):
    """休憩開始時刻 → "休憩 HH:MM-HH:MM" （1h固定）"""
    if not break_start:
        return ""
    h, m = break_start.split(":")
    end_h = int(h) + 1
    return f"休憩 {break_start}-{end_h:02d}:{m}"

# ③ 相談員事務スロットラベル
DESK_SLOT_LABELS = ["9-11時", "11-13時", "13-15時", "15-17時"]

# 休憩なし明示表示の対象パターン（半日・訪問のみ）
_NO_BREAK_PATTERNS = {"day_pattern3", "day_pattern4", "visit_am", "visit_pm",
                      "day_am", "day_pm"}  # 旧名の後方互換

# 兼務パターンは休憩はあるが相談業務なし
_NO_COUNSELOR_PATTERNS = {"day_p3_visit_pm", "visit_am_day_p4"}

# デイ午前に寄与するアサインメント
_DAY_AM_SET = {"day_pattern1", "day_pattern2", "day_pattern3", "day_p3_visit_pm",
               "day_am", "day_am_visit_pm"}
# デイ午後に寄与するアサインメント
_DAY_PM_SET = {"day_pattern1", "day_pattern2", "day_pattern4", "visit_am_day_p4",
               "day_pm", "visit_am_day_pm"}
# 訪問午前
_VISIT_AM_SET = {"visit_am", "visit_am_day_p4", "visit_am_day_pm"}
# 訪問午後
_VISIT_PM_SET = {"visit_pm", "day_p3_visit_pm", "day_am_visit_pm"}
# 兼務
_DUAL_SET = {"day_p3_visit_pm", "visit_am_day_p4", "day_am_visit_pm", "visit_am_day_pm"}
# 調理
_COOK_SET = {"cook_early", "cook_morning", "cook_late", "cook_long"}
_NURSE_PT_NAME_ALIASES = {"看護師", "PT", "理学療法士"}
_NURSE_PT_CODE_ALIASES = {"nurse", "pt"}


def _is_nurse_or_pt_staff(staff: dict) -> bool:
    """看護師/PTはコード優先で判定し、旧名称も受け入れる。"""
    qual_codes = {
        code for code in staff.get("qualification_codes", [])
        if isinstance(code, str) and code
    }
    if qual_codes.intersection(_NURSE_PT_CODE_ALIASES):
        return True

    qual_names = {
        name for name in staff.get("qualifications", [])
        if isinstance(name, str) and name
    }
    return bool(qual_names.intersection(_NURSE_PT_NAME_ALIASES))


# ---------------------------------------------------------------------------
# ヘルパー: 日ごとの配置データを集計する
# ---------------------------------------------------------------------------
def _build_daily_data(shifts_data, staff_list, year, month):
    """
    日付別・職員別の配置マップと、日付別サマリーを構築する。
    """
    num_days = calendar.monthrange(year, month)[1]
    dates = [date(year, month, d) for d in range(1, num_days + 1)]

    assignment_map = {}
    phone_duty_map = {}
    desk_slot_map = {}  # ③ {date_str: {staff_id: [slot_idx, ...]}}
    break_map = {}      # ① {date_str: {staff_id: "12:00"}}
    for item in shifts_data:
        d_str = item["date"]
        sid = item["staff_id"]
        asgn = item.get("assignment", "")
        if d_str not in assignment_map:
            assignment_map[d_str] = {}
        assignment_map[d_str][sid] = asgn
        if item.get("is_phone_duty"):
            if d_str not in phone_duty_map:
                phone_duty_map[d_str] = []
            phone_duty_map[d_str].append(item.get("staff_name", f"ID:{sid}"))
        # ③ 相談員事務スロット
        slots = item.get("counselor_desk_slots")
        if slots:
            if d_str not in desk_slot_map:
                desk_slot_map[d_str] = {}
            desk_slot_map[d_str][sid] = slots
        # ① 休憩開始時刻
        bs = item.get("break_start")
        if bs:
            if d_str not in break_map:
                break_map[d_str] = {}
            break_map[d_str][sid] = bs

    # ② 看護師/PTはデイ人数カウントから除外
    nurse_pt_ids = set()
    for st in staff_list:
        if _is_nurse_or_pt_staff(st):
            nurse_pt_ids.add(st["id"])

    summary_map = {}
    for d in dates:
        d_str = d.isoformat()
        day_assignments = assignment_map.get(d_str, {})

        day_am = 0
        day_pm = 0
        visit_am = 0
        visit_pm = 0
        dual = 0
        cook_total = 0

        for sid, asgn in day_assignments.items():
            is_nurse_pt = sid in nurse_pt_ids
            if asgn in _DAY_AM_SET and not is_nurse_pt:
                day_am += 1
            if asgn in _DAY_PM_SET and not is_nurse_pt:
                day_pm += 1
            if asgn in _VISIT_AM_SET:
                visit_am += 1
            if asgn in _VISIT_PM_SET:
                visit_pm += 1
            if asgn in _DUAL_SET:
                dual += 1
            if asgn in _COOK_SET:
                cook_total += 1

        summary_map[d_str] = {
            "day_am": day_am,
            "day_pm": day_pm,
            "visit_am": visit_am,
            "visit_pm": visit_pm,
            "dual": dual,
            "cook_total": cook_total,
        }

    return dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map


# ---------------------------------------------------------------------------
# Excel エクスポート
# ---------------------------------------------------------------------------
def export_excel(
    shifts_data: list,
    warnings_data: list,
    staff_list: list,
    year: int,
    month: int,
) -> BytesIO:
    """Excel 形式でシフト表を出力する。"""
    wb = Workbook()

    ws = wb.active
    ws.title = "シフト表"

    dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )

    care_staff = [s for s in staff_list if s.get("department") != "cooking"]
    cook_staff = [s for s in staff_list if s.get("department") == "cooking"]
    care_names = [s["name"] for s in care_staff]
    care_ids = [s["id"] for s in care_staff]
    cook_names = [s["name"] for s in cook_staff]
    cook_ids = [s["id"] for s in cook_staff]
    has_cooking = len(cook_staff) > 0

    care_summary_start_col = 3 + len(care_staff)
    cook_staff_start_col = care_summary_start_col + len(SUMMARY_HEADERS)

    if has_cooking:
        cook_summary_start_col = cook_staff_start_col + len(cook_staff)
        last_col = cook_summary_start_col + len(COOK_SUMMARY_HEADERS) - 1
    else:
        last_col = care_summary_start_col + len(SUMMARY_HEADERS) - 1

    # --- タイトル行 ---
    title_text = f"{year}年{month}月 シフト表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- ヘッダー行 ④ 資格表示付き ---
    header_row = 2
    # ④ ケアスタッフ名に資格情報を付加
    care_header_names = []
    for s in care_staff:
        quals = s.get("qualifications", [])
        if quals:
            care_header_names.append(f"{s['name']}\n({'/'.join(quals)})")
        else:
            care_header_names.append(s["name"])

    headers = ["日付", "曜日"] + care_header_names + SUMMARY_HEADERS
    if has_cooking:
        headers += cook_names + COOK_SUMMARY_HEADERS

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # --- データ行 ---
    data_start_row = 3

    for row_offset, d in enumerate(dates):
        row = data_start_row + row_offset
        d_str = d.isoformat()
        weekday_idx = d.weekday()
        weekday_name = WEEKDAY_NAMES[weekday_idx]

        # ⑧ 祝日判定
        is_holiday = jpholiday.is_holiday(d)
        row_fill = None
        if is_holiday:
            row_fill = HOLIDAY_FILL
        elif weekday_idx == 5:
            row_fill = SATURDAY_FILL
        elif weekday_idx == 6:
            row_fill = SUNDAY_FILL

        date_cell = ws.cell(row=row, column=1, value=f"{d.month}/{d.day}")
        date_cell.font = NORMAL_FONT
        date_cell.alignment = CENTER_ALIGN
        date_cell.border = THIN_BORDER
        if row_fill:
            date_cell.fill = row_fill

        # 祝日名を曜日セルに追記
        if is_holiday:
            holiday_name = jpholiday.is_holiday_name(d)
            dow_value = f"{weekday_name}\n{holiday_name}"
        else:
            dow_value = weekday_name
        dow_cell = ws.cell(row=row, column=2, value=dow_value)
        dow_cell.font = NORMAL_FONT
        dow_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dow_cell.border = THIN_BORDER
        if row_fill:
            dow_cell.fill = row_fill

        day_assignments = assignment_map.get(d_str, {})
        max_lines = 1
        for staff_offset, sid in enumerate(care_ids):
            col = 3 + staff_offset
            asgn = day_assignments.get(sid, "")
            label = ASSIGNMENT_LABELS.get(asgn, "")

            # ① 休憩・③ 相談をセル内テキストに追記（印刷対応）
            display_text = label
            lines = 1
            bs = break_map.get(d_str, {}).get(sid)
            break_text = _format_break_comment(bs)
            if break_text:
                display_text += f"\n{break_text}"
                lines += 1
            elif asgn in _NO_BREAK_PATTERNS:
                display_text += "\n休憩なし"
                lines += 1

            desk_slots = desk_slot_map.get(d_str, {}).get(sid)
            if desk_slots:
                slot_texts = [DESK_SLOT_LABELS[si] for si in desk_slots if si < len(DESK_SLOT_LABELS)]
                if slot_texts:
                    display_text += f"\n相談:{','.join(slot_texts)}"
                    lines += 1
            elif asgn in _NO_BREAK_PATTERNS or asgn in _NO_COUNSELOR_PATTERNS:
                display_text += "\n相談なし"
                lines += 1

            cell = ws.cell(row=row, column=col, value=display_text)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            if asgn in ASSIGNMENT_FILL:
                cell.fill = ASSIGNMENT_FILL[asgn]
            elif row_fill:
                cell.fill = row_fill

            if lines > max_lines:
                max_lines = lines

        summary = summary_map.get(d_str, {
            "day_am": 0, "day_pm": 0,
            "visit_am": 0, "visit_pm": 0, "dual": 0,
            "cook_total": 0,
        })
        phone_names = phone_duty_map.get(d_str, [])
        summary_values = [
            summary["day_am"],
            summary["day_pm"],
            summary["visit_am"],
            summary["visit_pm"],
            summary["dual"],
            ", ".join(phone_names) if phone_names else "",
        ]

        _SUMMARY_WARNING_TYPES = [
            'understaffed_day_am',
            'understaffed_day_pm',
            'understaffed_visit_am',
            'understaffed_visit_pm',
            'dual_shortage',
            None,
        ]

        for s_offset, val in enumerate(summary_values):
            col = care_summary_start_col + s_offset
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

            expected_type = _SUMMARY_WARNING_TYPES[s_offset] if s_offset < len(_SUMMARY_WARNING_TYPES) else None
            is_alert = False
            if expected_type:
                for w in warnings_data:
                    if w.get("date") == d_str and w.get("warning_type") == expected_type:
                        is_alert = True
                        break

            if is_alert:
                cell.fill = ALERT_FILL
                cell.font = ALERT_FONT
            elif row_fill:
                cell.fill = row_fill

        if has_cooking:
            for staff_offset, sid in enumerate(cook_ids):
                col = cook_staff_start_col + staff_offset
                asgn = day_assignments.get(sid, "")
                label = ASSIGNMENT_LABELS.get(asgn, "")

                # ① 休憩をセル内テキストに追記（印刷対応）
                display_text = label
                bs = break_map.get(d_str, {}).get(sid)
                break_text = _format_break_comment(bs)
                if break_text:
                    display_text += f"\n{break_text}"
                    if 2 > max_lines:
                        max_lines = 2

                cell = ws.cell(row=row, column=col, value=display_text)
                cell.font = NORMAL_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER

                if asgn in ASSIGNMENT_FILL:
                    cell.fill = ASSIGNMENT_FILL[asgn]
                elif row_fill:
                    cell.fill = row_fill

            cook_total = summary.get("cook_total", 0)
            col = cook_summary_start_col
            cell = ws.cell(row=row, column=col, value=cook_total)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

            is_cook_alert = False
            for w in warnings_data:
                if w.get("date") == d_str and w.get("warning_type", "").startswith("understaffed_cook"):
                    is_cook_alert = True
                    break
            if is_cook_alert:
                cell.fill = ALERT_FILL
                cell.font = ALERT_FONT
            elif row_fill:
                cell.fill = row_fill

        # 行高さを内容に応じて調整
        if max_lines >= 3:
            ws.row_dimensions[row].height = 45
        elif max_lines == 2:
            ws.row_dimensions[row].height = 30

    # --- ⑦ フッター: 出勤日数 ---
    footer_row = data_start_row + len(dates)
    footer_font = Font(name="メイリオ", bold=True, size=10)

    ws.cell(row=footer_row, column=1, value="出勤日数").font = footer_font
    ws.cell(row=footer_row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=footer_row, column=1).border = THIN_BORDER
    ws.cell(row=footer_row, column=2, value="").border = THIN_BORDER

    for staff_offset, sid in enumerate(care_ids):
        count = sum(
            1 for d in dates
            if assignment_map.get(d.isoformat(), {}).get(sid, "off") not in ("off", "")
        )
        cell = ws.cell(row=footer_row, column=3 + staff_offset, value=count)
        cell.font = footer_font
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # サマリー列は空セル
    for s_offset in range(len(SUMMARY_HEADERS)):
        cell = ws.cell(row=footer_row, column=care_summary_start_col + s_offset, value="")
        cell.border = THIN_BORDER

    if has_cooking:
        for staff_offset, sid in enumerate(cook_ids):
            count = sum(
                1 for d in dates
                if assignment_map.get(d.isoformat(), {}).get(sid, "cook_off") not in ("cook_off", "")
            )
            cell = ws.cell(row=footer_row, column=cook_staff_start_col + staff_offset, value=count)
            cell.font = footer_font
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for s_offset in range(len(COOK_SUMMARY_HEADERS)):
            cell = ws.cell(row=footer_row, column=cook_summary_start_col + s_offset, value="")
            cell.border = THIN_BORDER

    # --- 列幅 ---
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    for i in range(len(care_staff)):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 16
    for i in range(len(SUMMARY_HEADERS)):
        col_letter = get_column_letter(care_summary_start_col + i)
        ws.column_dimensions[col_letter].width = 10
    if has_cooking:
        for i in range(len(cook_staff)):
            col_letter = get_column_letter(cook_staff_start_col + i)
            ws.column_dimensions[col_letter].width = 12
        for i in range(len(COOK_SUMMARY_HEADERS)):
            col_letter = get_column_letter(cook_summary_start_col + i)
            ws.column_dimensions[col_letter].width = 10

    # --- 警告シート ---
    if warnings_data:
        ws_warn = wb.create_sheet(title="警告一覧")

        warn_headers = ["日付", "種別", "内容"]
        for col_idx, header in enumerate(warn_headers, start=1):
            cell = ws_warn.cell(row=1, column=col_idx, value=header)
            cell.font = WARNING_HEADER_FONT
            cell.fill = WARNING_HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        for row_offset, warn in enumerate(warnings_data):
            row = 2 + row_offset
            ws_warn.cell(row=row, column=1, value=warn.get("date", "")).font = NORMAL_FONT
            ws_warn.cell(row=row, column=1).border = THIN_BORDER
            ws_warn.cell(row=row, column=1).alignment = CENTER_ALIGN

            ws_warn.cell(
                row=row, column=2, value=warn.get("warning_type", "")
            ).font = NORMAL_FONT
            ws_warn.cell(row=row, column=2).border = THIN_BORDER
            ws_warn.cell(row=row, column=2).alignment = CENTER_ALIGN

            ws_warn.cell(
                row=row, column=3, value=warn.get("message", "")
            ).font = NORMAL_FONT
            ws_warn.cell(row=row, column=3).border = THIN_BORDER

        ws_warn.column_dimensions["A"].width = 12
        ws_warn.column_dimensions["B"].width = 18
        ws_warn.column_dimensions["C"].width = 50

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# CSV エクスポート
# ---------------------------------------------------------------------------
def export_csv(
    shifts_data: list,
    warnings_data: list,
    staff_list: list,
    year: int,
    month: int,
) -> str:
    """CSV 形式でシフト表を出力する。"""
    dates, assignment_map, summary_map, phone_duty_map, desk_slot_map, break_map = _build_daily_data(
        shifts_data, staff_list, year, month
    )

    care_staff = [s for s in staff_list if s.get("department") != "cooking"]
    cook_staff = [s for s in staff_list if s.get("department") == "cooking"]
    care_names = [s["name"] for s in care_staff]
    care_ids = [s["id"] for s in care_staff]
    cook_names = [s["name"] for s in cook_staff]
    cook_ids = [s["id"] for s in cook_staff]
    has_cooking = len(cook_staff) > 0

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["日付", "曜日"] + care_names + SUMMARY_HEADERS
    if has_cooking:
        headers += cook_names + COOK_SUMMARY_HEADERS
    writer.writerow(headers)

    for d in dates:
        d_str = d.isoformat()
        weekday_name = WEEKDAY_NAMES[d.weekday()]
        day_assignments = assignment_map.get(d_str, {})

        care_cells = []
        for sid in care_ids:
            asgn = day_assignments.get(sid, "")
            label = ASSIGNMENT_LABELS.get(asgn, "")
            care_cells.append(label)

        summary = summary_map.get(d_str, {
            "day_am": 0, "day_pm": 0,
            "visit_am": 0, "visit_pm": 0, "dual": 0,
            "cook_total": 0,
        })
        phone_names = phone_duty_map.get(d_str, [])
        summary_cells = [
            summary["day_am"],
            summary["day_pm"],
            summary["visit_am"],
            summary["visit_pm"],
            summary["dual"],
            ", ".join(phone_names) if phone_names else "",
        ]

        row = [f"{d.month}/{d.day}", weekday_name] + care_cells + summary_cells

        if has_cooking:
            cook_cells = []
            for sid in cook_ids:
                asgn = day_assignments.get(sid, "")
                label = ASSIGNMENT_LABELS.get(asgn, "")
                cook_cells.append(label)
            row += cook_cells + [summary.get("cook_total", 0)]

        writer.writerow(row)

    csv_string = "\ufeff" + output.getvalue()
    return csv_string
